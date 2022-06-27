from re import M
from openpyxl import Workbook, load_workbook
from datetime import date
def getRange(month:int)->int:
    if month in [1,3,5,7,8,10,12]:
        return 31
    elif month in  [4,6,9,11]:
        return 30
    else:
        return 28

wb = load_workbook('schedule.xlsx')
if 'google_calendar' in wb.sheetnames:  
    pass
else:
    wb.create_sheet('google_calendar')
    title=['Subject','Start Date','Start Time','End Date','End Time','All Day Event','Description','Location','Private']
    ws = wb['google_calendar']
    ws.append(title)
ws=wb.worksheets[0]
ws_c=wb.worksheets[1]
year = date.today().year
month = ws['A1']
date=getRange(month.value)
name=ws.cell(2,1)
# for col in ws.iter_cols(min_row=1,min_col=3,max_row=2,max_col=date+2,values_only=True):
#     for cell in col:
#         print(cell,end=' ')

# for i in range(2,ws.max_row+1):
#     for j in range(3,ws.max_column+1):
#         event=[None]*9
#         content=ws.cell(row=i,column=j).value
#         if content=='休':
#             event[0]=content
#             event[1]=str(year)+'/'+str(month.value)+'/'+str(j-2) 
#             event[3]=str(year)+'/'+str(month.value)+'/'+str(j-2)       
#             event[5]=True
#             event[8]=True
#         elif content=='早':
#             event[0]=content
#             event[1]=str(year)+'/'+str(month.value)+'/'+str(j-2)
#             event[2]='7:00'
#             event[3]=str(year)+'/'+str(month.value)+'/'+str(j-2)  
#             event[4]='16:00'     
#             event[5]=True
#             event[8]=True
#         elif content=='中':
#             event[0]=content
#             event[1]=str(year)+'/'+str(month.value)+'/'+str(j-2) 
#             event[2]='12:00'
#             event[3]=str(year)+'/'+str(month.value)+'/'+str(j-2)  
#             event[4]='21:00'     
#             event[5]=True
#             event[8]=True
#         elif content=='晚':
#             event[0]=content
#             event[1]=str(year)+'/'+str(month.value)+'/'+str(j-2) 
#             event[2]='14:00'
#             event[3]=str(year)+'/'+str(month.value)+'/'+str(j-2)  
#             event[4]='23:00'     
#             event[5]=True
#             event[8]=True
        
#         ws_c.append(event)
        
#     # print()
# wb.save('schedule.xlsx')

csv=open('data.csv','w+')
for row in range(1,ws_c.max_row):
    
    for column in range(1,ws_c.max_column):
        # print(ws_c.cell(row=row,column=column).value,end=' ')
        if column==ws_c.max_column-1:
            csv.write(str(ws_c.cell(row=row,column=column).value))
        else:
            csv.write(str(ws_c.cell(row,column).value)+',')
    csv.write('\n')
    
csv.close()

