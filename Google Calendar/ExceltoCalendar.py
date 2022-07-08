from re import M
from openpyxl import Workbook, load_workbook
from datetime import date
def CalendarFormat():
    def getRange(month:int,year)->int:
        if month in [1,3,5,7,8,10,12]:
            return 31
        elif month in  [4,6,9,11]:
            return 30
        else:
            if (year % 4==0 and year % 100!=0) or year %400==0:
                return 29
            return 28
                

    wb = load_workbook('./Google Calendar/schedule.xlsx')
    if 'google_calendar' in wb.sheetnames:  
        return
    else:
        print('create sheet')
        wb.create_sheet('google_calendar')
        title=['Subject','Start Date','Start Time','End Date','End Time','All Day Event','Description','Location','Private']
        ws = wb['google_calendar']
        if ws.cell(1,1).value is not None:
            
            print(ws.cell(1,1))
            return
        else:
           for i in range(1,10):
                ws.cell(row=1,column=i).value=title[i-1]
                print(ws.cell(row=1,column=i).value)
    # print(wb.sheetnames)
        ws=wb.worksheets[0]
        ws_c=wb.worksheets[1]
        year = date.today().year
        month = ws['A1']
        # date=getRange(month.value,year)
        name=ws.cell(2,1)

        #加入sheet
        
        for i in range(2,ws.max_row+1):
            for j in range(3,ws.max_column+1):
                event=[None]*9
                content=ws.cell(row=i,column=j).value
                if content=='休':
                    event[0]=content
                    event[1]=str(year)+'/'+str(month.value)+'/'+str(j-2) 
                    event[3]=str(year)+'/'+str(month.value)+'/'+str(j-2)       
                    event[5]=True
                    event[8]=True
                elif content=='早':
                    event[0]=content
                    event[1]=str(year)+'/'+str(month.value)+'/'+str(j-2)
                    event[2]='7:00'
                    event[3]=str(year)+'/'+str(month.value)+'/'+str(j-2)  
                    event[4]='16:00'     
                    event[5]=True
                    event[8]=True
                elif content=='中':
                    event[0]=content
                    event[1]=str(year)+'/'+str(month.value)+'/'+str(j-2) 
                    event[2]='12:00'
                    event[3]=str(year)+'/'+str(month.value)+'/'+str(j-2)  
                    event[4]='21:00'     
                    event[5]=True
                    event[8]=True
                elif content=='晚':
                    event[0]=content
                    event[1]=str(year)+'/'+str(month.value)+'/'+str(j-2) 
                    event[2]='14:00'
                    event[3]=str(year)+'/'+str(month.value)+'/'+str(j-2)  
                    event[4]='23:00'     
                    event[5]=True
                    event[8]=True
                
                ws_c.append(event)
                
            # print()
        wb.save('.\Google Calendar\schedule.xlsx')

def toCSV(name):
    wb = load_workbook('schedule.xlsx')
    ws_c=wb.worksheets[1]
    csv=open(name,'w+')
    for row in range(1,ws_c.max_row):
        
        for column in range(1,ws_c.max_column):
            # print(ws_c.cell(row=row,column=column).value,end=' ')
            if column==ws_c.max_column-1:
                csv.write(str(ws_c.cell(row=row,column=column).value))
            else:
                csv.write(str(ws_c.cell(row,column).value)+',')
        csv.write('\n')
        
    csv.close()


CalendarFormat()
